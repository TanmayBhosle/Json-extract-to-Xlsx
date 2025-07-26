import tkinter as tk
from tkinter import filedialog
import json
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

def normalize_name(name):
    """Replace invalid filename characters and strip whitespace."""
    return re.sub(r'[<>:"/\\|?*]', '_', name).strip()

def extract_requests(items, path=None):
    """Recursively extract requests and folder hierarchy."""
    if path is None:
        path = []
    results = []
    for item in items:
        if "item" in item:
            folder_name = item.get("name", "").strip()
            split_folders = [normalize_name(part) for part in folder_name.split("/") if part]
            results.extend(extract_requests(item["item"], path + split_folders))
        elif "request" in item:
            req_name = normalize_name(item.get("name", "Unnamed Request"))
            method = item["request"].get("method", "GET")
            url_data = item["request"].get("url", {})
            raw_url = ""

            if isinstance(url_data, dict):
                raw_url = url_data.get("raw", "")
                if not raw_url:
                    protocol = url_data.get("protocol", "https")
                    host = ".".join(url_data.get("host", [])) if "host" in url_data else ""
                    path_parts = "/".join(url_data.get("path", [])) if "path" in url_data else ""
                    raw_url = f"{protocol}://{host}/{path_parts}".rstrip("/")
            elif isinstance(url_data, str):
                raw_url = url_data

            results.append(path + [req_name, method, raw_url])
    return results

def auto_adjust_column_width(ws):
    """Auto-fit column width based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

def main():
    # Hide main Tk window
    root = tk.Tk()
    root.withdraw()

    # Ask user to select Postman collection JSON file
    input_file = filedialog.askopenfilename(
        title="Select Postman Collection JSON",
        filetypes=[("JSON Files", "*.json")]
    )

    if not input_file:
        print("❌ No input file selected.")
        return
    if not os.path.isfile(input_file):
        print(f"❌ File not found: {input_file}")
        return

    print(f"📂 Selected JSON file: {input_file}")

    # Load JSON data
    try:
        with open(input_file, "r", encoding="utf-8") as f:
            collection = json.load(f)
    except Exception as e:
        print(f"❌ Error reading JSON: {e}")
        return

    # Extract requests
    all_requests = extract_requests(collection.get("item", []))
    if not all_requests:
        print("⚠️ No requests found in the JSON file.")
        return

    # Print sample output
    print("\n🔍 Sample extracted requests:")
    for row in all_requests[:5]:
        print(row)

    # Determine folder depth for column generation
    max_folder_depth = max(len(row) - 3 for row in all_requests)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Postman URLs"

    headers = [f"Folder Level {i+1}" for i in range(max_folder_depth)] + ["Request Name", "Method", "URL"]
    ws.append(headers)

    for row in all_requests:
        folders = row[:-3]
        folders += [""] * (max_folder_depth - len(folders))
        req_name, method, url = row[-3:]
        ws.append(folders + [req_name, method, url])

    auto_adjust_column_width(ws)

    # Ask user where to save the Excel file
    output_file = filedialog.asksaveasfilename(
        title="Save Excel File As",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        initialfile="postman_urls_separated_folders.xlsx"
    )

    if not output_file:
        print("❌ No output file selected. Operation cancelled.")
        return

    try:
        wb.save(output_file)
        print(f"\n✅ Excel file saved successfully: {output_file}")
    except Exception as e:
        print(f"❌ Failed to save Excel file: {e}")

if __name__ == "__main__":
    main()
